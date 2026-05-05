"""Import transaksi + nasabah DLH ke database backend.

Usage examples:
    python import_dlh_excel.py --limit 100
    python import_dlh_excel.py --commit
    python import_dlh_excel.py --transaction-sheet Sheet1 --customer-sheet Nasabah --commit
"""

from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import text

from app import create_app, db
from app.models.user import User
from app.models.waste_deposit import WasteDeposit
from app.models.waste_point_rate import WastePointRate
from app.services.gamification_service import (
    calculate_user_level,
    ensure_point_settings_seeded,
    ensure_waste_point_rates_seeded,
)


TX_ALIASES = {
    "timestamp": ["Timestamp"],
    "activity_type": ["Jenis Kegiatan"],
    "name": ["Nama Nasabah", "Nama"],
    "account_number": ["Nomor Rekening", "No Rekening", "No Rekening "],
    "waste": ["Jenis"],
    "weight": ["Berat (kg)"],
    "price": ["Harga (Rp.) / Kg", "Harga (Rp.)/Kg", "Harga / Kg"],
    "total": ["Total Tabungan (Rp.)", "Total Tabungan"],
}

CUSTOMER_ALIASES = {
    "timestamp": ["Timestamp"],
    "account_number": ["No Rekening", "No. Rekening", "Nomor Rekening"],
    "name": ["Nama", "Nama Nasabah", "Nama Lengkap"],
    "gender": ["Jenis Kelamin"],
    "nik": ["NIK"],
    "address": ["Alamat"],
    "department": ["Bidang"],
}


def _normalize_header(value: object) -> str:
    text_value = str(value or "").strip()
    return " ".join(text_value.split())


def _clean_account_number(value: object) -> str:
    return str(value or "").strip().replace(" ", "").upper()


def _safe_text(value: object) -> str | None:
    text_value = str(value or "").strip()
    return text_value if text_value else None


def _safe_float(raw: object) -> float | None:
    if raw is None or raw == "":
        return None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return None


def _safe_int(raw: object) -> int | None:
    value = _safe_float(raw)
    if value is None:
        return None
    return int(round(value))


def _as_utc(dt: datetime | None) -> datetime:
    if dt is None:
        return datetime.now(timezone.utc)
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _resolve_waste_type(raw_jenis: object, valid_codes: set[str], name_to_code: dict[str, str]) -> str | None:
    text_value = str(raw_jenis or "").strip()
    if not text_value:
        return None

    token = text_value.split("-", 1)[0].strip().upper()
    if token in valid_codes:
        return token.lower()

    up = text_value.upper()
    if up in valid_codes:
        return up.lower()

    by_name = name_to_code.get(text_value.lower())
    if by_name:
        return by_name.lower()

    return None


def _default_excel_path() -> Path:
    return Path(__file__).resolve().parent.parent / "data" / "Data DLH.xlsx"


def _header_map(ws) -> dict[str, int]:
    headers = [_normalize_header(c.value) for c in ws[1]]
    return {name: idx for idx, name in enumerate(headers) if name}


def _find_col(header_index: dict[str, int], aliases: list[str]) -> int | None:
    for alias in aliases:
        normalized = _normalize_header(alias)
        if normalized in header_index:
            return header_index[normalized]
    return None


def _detect_sheet_for_aliases(wb, aliases: dict[str, list[str]], required_keys: list[str]) -> str | None:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_index = _header_map(ws)
        found_all = True
        for key in required_keys:
            if _find_col(header_index, aliases[key]) is None:
                found_all = False
                break
        if found_all:
            return sheet_name
    return None


def _sync_schema_columns() -> None:
    """Ensure required columns exist in current DB, without manual migration step."""
    statements = [
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS gender VARCHAR(20)",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS nik VARCHAR(32)",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS address VARCHAR(255)",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS department VARCHAR(120)",
        "CREATE INDEX IF NOT EXISTS ix_users_nik ON users (nik)",
        "ALTER TABLE waste_deposits ADD COLUMN IF NOT EXISTS activity_type VARCHAR(100)",
        "ALTER TABLE waste_deposits ADD COLUMN IF NOT EXISTS source_waste_label VARCHAR(120)",
        "ALTER TABLE waste_deposits ADD COLUMN IF NOT EXISTS source_price_per_kg INTEGER",
        "ALTER TABLE waste_deposits ADD COLUMN IF NOT EXISTS source_total_savings INTEGER",
    ]
    for stmt in statements:
        db.session.execute(text(stmt))
    db.session.commit()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import DLH transactions and customer profile data")
    parser.add_argument("--file", default=str(_default_excel_path()), help="Path to .xlsx file")
    parser.add_argument("--customer-file", default=None, help="Path file nasabah .xlsx (opsional)")
    parser.add_argument("--transaction-sheet", default=None, help="Sheet transaksi (auto-detect jika kosong)")
    parser.add_argument("--customer-sheet", default=None, help="Sheet nasabah (opsional, auto-detect jika kosong)")
    parser.add_argument("--admin-email", default="admin@dlh.padang.go.id", help="Validator admin email")
    parser.add_argument("--default-password", default="import123", help="Password user baru")
    parser.add_argument("--limit", type=int, default=0, help="Maksimum baris transaksi diproses (0=semua)")
    parser.add_argument("--batch-size", type=int, default=2000, help="Ukuran batch commit saat --commit")
    parser.add_argument("--check-duplicates", action="store_true", help="Cek duplikat deposit sebelum insert")
    parser.add_argument("--commit", action="store_true", help="Simpan ke database")
    return parser.parse_args()


def import_dlh_data(
    *,
    file: str,
    customer_file: str | None = None,
    transaction_sheet: str | None = None,
    customer_sheet: str | None = None,
    admin_email: str = "admin@dlh.padang.go.id",
    default_password: str = "import123",
    limit: int = 0,
    batch_size: int = 2000,
    check_duplicates: bool = False,
    commit: bool = True,
) -> dict:
    _sync_schema_columns()
    ensure_waste_point_rates_seeded()
    ensure_point_settings_seeded()

    admin = User.query.filter_by(email=admin_email).first()
    admin_id = admin.id if admin else None

    rates = WastePointRate.query.filter_by(is_active=True).all()
    valid_codes = {r.code.upper() for r in rates}
    name_to_code = {r.name.lower(): r.code.upper() for r in rates}
    points_per_code = {r.code.upper(): int(r.points_per_kg) for r in rates}

    file_path = Path(file)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    customer_file_path = Path(customer_file) if customer_file else file_path
    if not customer_file_path.exists():
        raise FileNotFoundError(f"Customer file not found: {customer_file_path}")
    customer_wb = wb if customer_file_path == file_path else load_workbook(customer_file_path, data_only=True)

    tx_sheet_name = transaction_sheet or _detect_sheet_for_aliases(
        wb,
        TX_ALIASES,
        required_keys=["timestamp", "account_number", "waste", "weight"],
    )
    if not tx_sheet_name:
        raise ValueError("Tidak dapat menemukan sheet transaksi secara otomatis.")

    customer_sheet_name = customer_sheet or _detect_sheet_for_aliases(
        customer_wb,
        CUSTOMER_ALIASES,
        required_keys=["account_number", "name", "nik"],
    )

    tx_ws = wb[tx_sheet_name]
    tx_header = _header_map(tx_ws)

    tx_cols = {
        key: _find_col(tx_header, aliases)
        for key, aliases in TX_ALIASES.items()
    }
    required_tx_keys = ["timestamp", "account_number", "waste", "weight"]
    missing_tx = [k for k in required_tx_keys if tx_cols[k] is None]
    if missing_tx:
        raise ValueError(f"Kolom transaksi wajib tidak ditemukan: {missing_tx}")

    customer_map: dict[str, dict[str, object]] = {}
    if customer_sheet_name:
        customer_ws = customer_wb[customer_sheet_name]
        customer_header = _header_map(customer_ws)
        customer_cols = {
            key: _find_col(customer_header, aliases)
            for key, aliases in CUSTOMER_ALIASES.items()
        }
        if customer_cols["account_number"] is not None:
            for row in customer_ws.iter_rows(min_row=2, values_only=True):
                account_number = _clean_account_number(row[customer_cols["account_number"]])
                if not account_number:
                    continue
                customer_map[account_number] = {
                    "name": _safe_text(row[customer_cols["name"]]) if customer_cols["name"] is not None else None,
                    "gender": _safe_text(row[customer_cols["gender"]]) if customer_cols["gender"] is not None else None,
                    "nik": _safe_text(row[customer_cols["nik"]]) if customer_cols["nik"] is not None else None,
                    "address": _safe_text(row[customer_cols["address"]]) if customer_cols["address"] is not None else None,
                    "department": _safe_text(row[customer_cols["department"]]) if customer_cols["department"] is not None else None,
                }

    stats = {
        "rows_seen": 0,
        "rows_imported": 0,
        "rows_skipped": 0,
        "users_created": 0,
        "users_updated": 0,
        "duplicates_skipped": 0,
        "invalid_waste_type": 0,
        "invalid_weight": 0,
        "missing_account": 0,
    }

    users_by_account: dict[str, User] = {}
    touched_user_ids: set[int] = set()

    for row in tx_ws.iter_rows(min_row=2, values_only=True):
        if limit and stats["rows_seen"] >= limit:
            break
        stats["rows_seen"] += 1

        account_number = _clean_account_number(row[tx_cols["account_number"]])
        if not account_number:
            stats["rows_skipped"] += 1
            stats["missing_account"] += 1
            continue

        weight_kg = _safe_float(row[tx_cols["weight"]])
        if weight_kg is None or weight_kg <= 0:
            stats["rows_skipped"] += 1
            stats["invalid_weight"] += 1
            continue

        raw_waste = row[tx_cols["waste"]]
        waste_type = _resolve_waste_type(raw_waste, valid_codes, name_to_code)
        if not waste_type:
            stats["rows_skipped"] += 1
            stats["invalid_waste_type"] += 1
            continue

        created_at_raw = row[tx_cols["timestamp"]]
        created_at = _as_utc(created_at_raw if isinstance(created_at_raw, datetime) else None)
        activity_type = _safe_text(row[tx_cols["activity_type"]]) if tx_cols["activity_type"] is not None else None
        source_waste_label = _safe_text(raw_waste)
        source_price_per_kg = _safe_int(row[tx_cols["price"]]) if tx_cols["price"] is not None else None
        source_total_savings = _safe_int(row[tx_cols["total"]]) if tx_cols["total"] is not None else None
        tx_name = _safe_text(row[tx_cols["name"]]) if tx_cols["name"] is not None else None

        user = users_by_account.get(account_number)
        if not user:
            user = User.query.filter_by(account_number=account_number).first()
            if not user:
                customer_info = customer_map.get(account_number, {})
                user_name = customer_info.get("name") or tx_name or account_number
                generated_email = f"{account_number.lower()}@import.dlh.local"
                user = User(
                    name=str(user_name),
                    email=generated_email,
                    account_number=account_number,
                    gender=customer_info.get("gender"),
                    nik=customer_info.get("nik"),
                    address=customer_info.get("address"),
                    department=customer_info.get("department"),
                    role="member",
                    level="Bronze",
                    total_points=0,
                )
                user.set_password(default_password)
                db.session.add(user)
                db.session.flush()
                stats["users_created"] += 1
            users_by_account[account_number] = user

        customer_info = customer_map.get(account_number)
        user_updated = False
        if customer_info:
            if customer_info.get("name") and user.name != customer_info["name"]:
                user.name = str(customer_info["name"])
                user_updated = True
            if customer_info.get("gender") and user.gender != customer_info["gender"]:
                user.gender = str(customer_info["gender"])
                user_updated = True
            if customer_info.get("nik") and user.nik != customer_info["nik"]:
                user.nik = str(customer_info["nik"])
                user_updated = True
            if customer_info.get("address") and user.address != customer_info["address"]:
                user.address = str(customer_info["address"])
                user_updated = True
            if customer_info.get("department") and user.department != customer_info["department"]:
                user.department = str(customer_info["department"])
                user_updated = True
        elif tx_name and user.name != tx_name:
            user.name = tx_name
            user_updated = True

        if user_updated:
            stats["users_updated"] += 1

        if check_duplicates:
            duplicate = WasteDeposit.query.filter_by(
                user_id=user.id,
                created_at=created_at,
                weight_kg=weight_kg,
                waste_type=waste_type,
            ).first()
            if duplicate:
                stats["rows_skipped"] += 1
                stats["duplicates_skipped"] += 1
                continue

        points_rate = points_per_code.get(waste_type.upper(), 0)
        points = int(weight_kg * points_rate)

        deposit = WasteDeposit(
            user_id=user.id,
            weight_kg=weight_kg,
            waste_type=waste_type,
            activity_type=activity_type,
            source_waste_label=source_waste_label,
            source_price_per_kg=source_price_per_kg,
            source_total_savings=source_total_savings,
            status="validated",
            points_earned=points,
            created_at=created_at,
            validated_at=created_at,
            validated_by=admin_id,
        )
        db.session.add(deposit)
        touched_user_ids.add(user.id)
        stats["rows_imported"] += 1

        if commit and batch_size > 0 and stats["rows_imported"] % batch_size == 0:
            db.session.commit()

    if touched_user_ids:
        totals = db.session.query(
            WasteDeposit.user_id,
            db.func.coalesce(db.func.sum(WasteDeposit.points_earned), 0),
        ).filter(
            WasteDeposit.user_id.in_(touched_user_ids),
            WasteDeposit.status == "validated",
        ).group_by(WasteDeposit.user_id).all()

        totals_map = {int(uid): int(total or 0) for uid, total in totals}
        for uid in touched_user_ids:
            user = db.session.get(User, uid)
            if not user:
                continue
            user.total_points = totals_map.get(uid, 0)
            user.level = calculate_user_level(user.total_points)

    if commit:
        db.session.commit()
        mode = "COMMIT"
    else:
        db.session.rollback()
        mode = "DRY-RUN (no changes saved)"

    return {
        "mode": mode,
        "source_file": str(file_path),
        "customer_file": str(customer_file_path),
        "transaction_sheet": tx_sheet_name,
        "customer_sheet": customer_sheet_name or "-",
        "stats": stats,
    }


def main() -> None:
    args = parse_args()
    app = create_app("production")

    with app.app_context():
        result = import_dlh_data(
            file=args.file,
            customer_file=args.customer_file,
            transaction_sheet=args.transaction_sheet,
            customer_sheet=args.customer_sheet,
            admin_email=args.admin_email,
            default_password=args.default_password,
            limit=args.limit,
            batch_size=args.batch_size,
            check_duplicates=args.check_duplicates,
            commit=args.commit,
        )

        print("=" * 60)
        print(f"Import mode: {result['mode']}")
        print(f"Source file: {result['source_file']}")
        print(f"Customer file: {result['customer_file']}")
        print(f"Transaction sheet: {result['transaction_sheet']}")
        print(f"Customer sheet: {result['customer_sheet']}")
        print("-" * 60)
        for key, value in result["stats"].items():
            print(f"{key}: {value}")
        print("=" * 60)


if __name__ == "__main__":
    main()
