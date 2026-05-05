from datetime import datetime, timezone, timedelta
import csv
import io
import os
import tempfile
import uuid
from flask import Blueprint, request, jsonify, current_app, send_from_directory, url_for
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import func
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from app import db
from app.models.user import User
from app.models.waste_deposit import WasteDeposit
from app.models.mission import Mission, UserMission
from app.models.badge import Badge, UserBadge
from app.models.reward import Reward, RewardRedemption
from app.models.participation_risk import ParticipationRisk
from app.models.waste_point_rate import WastePointRate
from app.models.point_setting import PointSetting
from app.services.gamification_service import (
    ensure_waste_point_rates_seeded,
    get_waste_display_name,
    ensure_point_settings_seeded,
    sync_all_users_levels_and_badges,
)
from app.services.simple_cache import invalidate_cache
from app.utils.api_response import error_response
from import_dlh_excel import import_dlh_data, CUSTOMER_ALIASES

admin_bp = Blueprint('admin', __name__)

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'webp'}
ALLOWED_IMAGE_MIME_TYPES = {'image/png', 'image/jpeg', 'image/webp'}


def _get_current_user():
    user_id = int(get_jwt_identity())
    return User.query.get(user_id)


def _require_admin(user):
    if not user or not user.is_admin:
        return error_response("Akses ditolak, hanya admin", "forbidden", status=403)
    return None


def _serialize_deposit_with_user(deposit):
    return {
        'id': str(deposit.id),
        'user_id': str(deposit.user_id),
        'user_name': deposit.user.name if deposit.user else '-',
        'account_number': deposit.user.account_number if deposit.user else '-',
        'weight_kg': float(deposit.weight_kg or 0),
        'waste_type': deposit.waste_type,
        'waste_label': get_waste_display_name(deposit.waste_type),
        'status': deposit.status,
        'points_earned': int(deposit.points_earned or 0),
        'created_at': deposit.created_at.isoformat() if deposit.created_at else None,
        'validated_at': deposit.validated_at.isoformat() if deposit.validated_at else None,
        'validated_by': str(deposit.validated_by) if deposit.validated_by else None,
        'notes': None,
    }


@admin_bp.route('/waste-point-rates', methods=['GET'])
@jwt_required()
def get_waste_point_rates():
    """Admin gets configurable waste point rates."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    ensure_waste_point_rates_seeded()
    rates = WastePointRate.query.order_by(WastePointRate.sort_order.asc(), WastePointRate.code.asc()).all()
    return jsonify({"rates": [r.to_dict() for r in rates]}), 200


@admin_bp.route('/waste-point-rates', methods=['POST'])
@jwt_required()
def create_waste_point_rate():
    """Admin adds a new waste type point rate."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    payload = request.get_json() or {}
    code = (payload.get('code') or '').strip().upper()
    name = (payload.get('name') or '').strip()
    category = (payload.get('category') or '').strip().lower() or 'lainnya'
    points = payload.get('points_per_kg')
    is_active = payload.get('is_active', True)

    if not code or not name or points is None:
        return error_response(
            "code, name, dan points_per_kg wajib diisi",
            "validation_error",
            status=400,
            fields={
                "code": "required" if not code else None,
                "name": "required" if not name else None,
                "points_per_kg": "required" if points is None else None,
            },
        )

    if len(code) > 10:
        return error_response("code maksimal 10 karakter", "validation_error", status=400, fields={"code": "max_length_10"})

    existing = WastePointRate.query.filter_by(code=code).first()
    if existing:
        return error_response("Code sampah sudah digunakan", "validation_error", status=400, fields={"code": "duplicate"})

    try:
        points_value = int(points)
    except (ValueError, TypeError):
        return error_response("points_per_kg harus angka", "validation_error", status=400, fields={"points_per_kg": "invalid"})

    if points_value < 0:
        return error_response("points_per_kg tidak boleh negatif", "validation_error", status=400, fields={"points_per_kg": "min_0"})

    max_sort = db.session.query(func.coalesce(func.max(WastePointRate.sort_order), 0)).scalar() or 0
    rate = WastePointRate(
        code=code,
        name=name,
        category=category,
        points_per_kg=points_value,
        is_active=bool(is_active),
        sort_order=int(max_sort) + 1,
    )

    db.session.add(rate)
    db.session.commit()

    invalidate_cache('waste_point_rates_active')

    return jsonify({
        "message": "Jenis sampah berhasil ditambahkan",
        "rate": rate.to_dict(),
    }), 201


@admin_bp.route('/waste-point-rates', methods=['PUT'])
@jwt_required()
def update_waste_point_rates():
    """Admin bulk updates waste point rates."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    payload = request.get_json() or {}
    items = payload.get('rates')
    if not isinstance(items, list) or not items:
        return error_response("rates wajib berupa list dan tidak boleh kosong", "validation_error", status=400, fields={"rates": "invalid"})

    ensure_waste_point_rates_seeded()

    for item in items:
        rate_id = item.get('id')
        code = (item.get('code') or '').strip().upper()
        name = (item.get('name') or '').strip()
        category = (item.get('category') or '').strip().lower()
        points = item.get('points_per_kg')
        is_active = item.get('is_active')

        if not rate_id or not code:
            return error_response(
                "id dan code wajib diisi",
                "validation_error",
                status=400,
                fields={
                    "id": "required" if not rate_id else None,
                    "code": "required" if not code else None,
                },
            )

        rate = WastePointRate.query.get(rate_id)
        if not rate:
            return error_response(f"Rate dengan id {rate_id} tidak ditemukan", "not_found", status=404)

        if points is None:
            return error_response(
                f"points_per_kg untuk {code} wajib diisi",
                "validation_error",
                status=400,
                fields={"points_per_kg": "required"},
            )

        if not name:
            return error_response(
                f"name untuk {code} wajib diisi",
                "validation_error",
                status=400,
                fields={"name": "required"},
            )

        try:
            points_value = int(points)
        except (ValueError, TypeError):
            return error_response(
                f"points_per_kg untuk {code} harus angka",
                "validation_error",
                status=400,
                fields={"points_per_kg": "invalid"},
            )

        if points_value < 0:
            return error_response(
                f"points_per_kg untuk {code} tidak boleh negatif",
                "validation_error",
                status=400,
                fields={"points_per_kg": "min_0"},
            )

        rate.points_per_kg = points_value
        rate.name = name
        if category:
            rate.category = category
        if isinstance(is_active, bool):
            rate.is_active = is_active

    db.session.commit()

    invalidate_cache('waste_point_rates_active')

    rates = WastePointRate.query.order_by(WastePointRate.sort_order.asc(), WastePointRate.code.asc()).all()
    return jsonify({
        "message": "Pengaturan poin sampah berhasil disimpan",
        "rates": [r.to_dict() for r in rates],
    }), 200


@admin_bp.route('/point-settings', methods=['GET'])
@jwt_required()
def get_point_settings():
    """Admin gets configurable level threshold point settings."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    ensure_point_settings_seeded()
    settings = PointSetting.query.order_by(PointSetting.sort_order.asc()).all()
    return jsonify({"settings": [s.to_dict() for s in settings]}), 200


@admin_bp.route('/point-settings', methods=['PUT'])
@jwt_required()
def update_point_settings():
    """Admin bulk updates level threshold point settings."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    payload = request.get_json() or {}
    items = payload.get('settings')
    if not isinstance(items, list) or not items:
        return error_response("settings wajib berupa list dan tidak boleh kosong", "validation_error", status=400, fields={"settings": "invalid"})

    ensure_point_settings_seeded()

    for item in items:
        setting_id = item.get('id')
        key = (item.get('key') or '').strip()
        name = (item.get('name') or '').strip()
        value = item.get('value')
        sort_order = item.get('sort_order')

        if not setting_id or not key:
            return error_response(
                "id dan key wajib diisi",
                "validation_error",
                status=400,
                fields={
                    "id": "required" if not setting_id else None,
                    "key": "required" if not key else None,
                },
            )

        setting = PointSetting.query.get(setting_id)
        if not setting:
            return error_response(f"Pengaturan poin dengan id {setting_id} tidak ditemukan", "not_found", status=404)

        if value is None:
            return error_response(
                f"value untuk {key} wajib diisi",
                "validation_error",
                status=400,
                fields={"value": "required"},
            )

        try:
            value_int = int(value)
        except (ValueError, TypeError):
            return error_response(
                f"value untuk {key} harus angka",
                "validation_error",
                status=400,
                fields={"value": "invalid"},
            )

        if value_int < 0:
            return error_response(
                f"value untuk {key} tidak boleh negatif",
                "validation_error",
                status=400,
                fields={"value": "min_0"},
            )

        if name:
            setting.name = name
        setting.value = value_int

        if sort_order is not None:
            try:
                setting.sort_order = int(sort_order)
            except (ValueError, TypeError):
                return error_response(
                    f"sort_order untuk {key} harus angka",
                    "validation_error",
                    status=400,
                    fields={"sort_order": "invalid"},
                )

    # Ensure thresholds are strictly increasing.
    latest = PointSetting.query.order_by(PointSetting.sort_order.asc()).all()
    threshold_values = [int(s.value) for s in latest]
    if threshold_values != sorted(threshold_values) or len(set(threshold_values)) != len(threshold_values):
        db.session.rollback()
        return error_response(
            "Nilai threshold level harus berurutan naik dan tidak boleh sama",
            "validation_error",
            status=400,
        )

    db.session.commit()

    invalidate_cache('point_settings')

    sync_stats = sync_all_users_levels_and_badges()
    level_badges = Badge.query.filter(Badge.name.like('Badge Level %')).order_by(Badge.condition_value.asc()).all()

    settings = PointSetting.query.order_by(PointSetting.sort_order.asc()).all()
    return jsonify({
        "message": "Pengaturan level poin berhasil disimpan",
        "settings": [s.to_dict() for s in settings],
        "level_badges": [b.to_dict() for b in level_badges],
        "sync": sync_stats,
    }), 200


@admin_bp.route('/sync-gamification', methods=['POST'])
@jwt_required()
def sync_gamification():
    """Manual trigger to sync level thresholds and badge assignments for all users."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    sync_stats = sync_all_users_levels_and_badges()
    level_badges = Badge.query.filter(Badge.name.like('Badge Level %')).order_by(Badge.condition_value.asc()).all()

    return jsonify({
        "message": "Sinkronisasi level dan badge berhasil dijalankan",
        "level_badges": [b.to_dict() for b in level_badges],
        "sync": sync_stats,
    }), 200


# ─── Data Management ──────────────────────────────────────────────────

@admin_bp.route('/data/import/members', methods=['POST'])
@jwt_required()
def import_members_from_file():
    """Admin imports member data from uploaded Excel file."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename:
        return error_response("File anggota wajib dipilih", "validation_error", status=400, fields={"file": "required"})

    try:
        wb = load_workbook(io.BytesIO(uploaded.read()), data_only=True)
    except Exception:
        return error_response("File anggota tidak valid. Gunakan format .xlsx", "validation_error", status=400, fields={"file": "invalid"})

    sheet_name = _detect_customer_sheet_name(wb)
    if not sheet_name:
        return error_response(
            "Sheet anggota tidak ditemukan. Pastikan ada kolom akun dan nama anggota",
            "validation_error",
            status=400,
            fields={"sheet": "missing"},
        )

    ws = wb[sheet_name]
    header_index = _header_map(ws)
    col_account = _find_col(header_index, CUSTOMER_ALIASES['account_number'])
    col_name = _find_col(header_index, CUSTOMER_ALIASES['name'])
    col_gender = _find_col(header_index, CUSTOMER_ALIASES['gender'])
    col_nik = _find_col(header_index, CUSTOMER_ALIASES['nik'])
    col_address = _find_col(header_index, CUSTOMER_ALIASES['address'])
    col_department = _find_col(header_index, CUSTOMER_ALIASES['department'])

    if col_account is None or col_name is None:
        return error_response(
            "Kolom wajib anggota tidak lengkap: No Rekening dan Nama",
            "validation_error",
            status=400,
            fields={"columns": "missing_required"},
        )

    stats = {
        'rows_seen': 0,
        'members_created': 0,
        'members_updated': 0,
        'rows_skipped': 0,
        'missing_account': 0,
        'missing_name': 0,
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats['rows_seen'] += 1

        account_number = _clean_account_number(row[col_account])
        if not account_number:
            stats['rows_skipped'] += 1
            stats['missing_account'] += 1
            continue

        name = _safe_text(row[col_name])
        if not name:
            stats['rows_skipped'] += 1
            stats['missing_name'] += 1
            continue

        gender = _safe_text(row[col_gender]) if col_gender is not None else None
        nik = _safe_text(row[col_nik]) if col_nik is not None else None
        address = _safe_text(row[col_address]) if col_address is not None else None
        department = _safe_text(row[col_department]) if col_department is not None else None

        member = User.query.filter_by(account_number=account_number).first()
        if not member:
            member = User(
                name=name,
                email=_generate_unique_import_email(account_number),
                account_number=account_number,
                gender=gender,
                nik=nik,
                address=address,
                department=department,
                role='member',
                level='Bronze',
                total_points=0,
            )
            member.set_password('import123')
            db.session.add(member)
            stats['members_created'] += 1
            continue

        updated = False
        if member.name != name:
            member.name = name
            updated = True
        if gender and member.gender != gender:
            member.gender = gender
            updated = True
        if nik and member.nik != nik:
            member.nik = nik
            updated = True
        if address and member.address != address:
            member.address = address
            updated = True
        if department and member.department != department:
            member.department = department
            updated = True

        if updated:
            stats['members_updated'] += 1

    db.session.commit()

    return jsonify({
        "message": "Import data anggota berhasil",
        "sheet": sheet_name,
        "stats": stats,
    }), 200


@admin_bp.route('/data/import/deposits', methods=['POST'])
@jwt_required()
def import_deposits_from_file():
    """Admin imports deposits data from uploaded Excel file."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename:
        return error_response("File setoran wajib dipilih", "validation_error", status=400, fields={"file": "required"})

    try:
        limit = int((request.form.get('limit') or '0').strip() or 0)
    except ValueError:
        return error_response("limit harus angka", "validation_error", status=400, fields={"limit": "invalid"})

    check_duplicates = str(request.form.get('check_duplicates', 'true')).lower() in ('1', 'true', 'yes', 'y')

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            uploaded.stream.seek(0)
            tmp.write(uploaded.stream.read())
            tmp_path = tmp.name

        result = import_dlh_data(
            file=tmp_path,
            customer_file=None,
            check_duplicates=check_duplicates,
            limit=max(0, limit),
            batch_size=2000,
            commit=True,
        )

        sync_stats = sync_all_users_levels_and_badges()

        return jsonify({
            "message": "Import data setoran berhasil",
            "import": result,
            "sync": sync_stats,
        }), 200
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)


@admin_bp.route('/data/import', methods=['POST'])
@jwt_required()
def import_data_from_excel():
    """Admin imports member and deposit data from standard DLH Excel files."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    payload = request.get_json() or {}
    check_duplicates = bool(payload.get('check_duplicates', True))
    limit = int(payload.get('limit', 0) or 0)
    batch_size = int(payload.get('batch_size', 2000) or 2000)

    data_dir = _data_dir_path()
    tx_file = os.path.join(data_dir, 'Data DLH.xlsx')
    customer_file = os.path.join(data_dir, 'data_nasabah.xlsx')

    if not os.path.exists(tx_file):
        return error_response(f"File transaksi tidak ditemukan: {tx_file}", "not_found", status=404)
    if not os.path.exists(customer_file):
        return error_response(f"File nasabah tidak ditemukan: {customer_file}", "not_found", status=404)

    result = import_dlh_data(
        file=tx_file,
        customer_file=customer_file,
        check_duplicates=check_duplicates,
        limit=limit,
        batch_size=batch_size,
        commit=True,
    )

    sync_stats = sync_all_users_levels_and_badges()

    return jsonify({
        "message": "Import data anggota dan setoran berhasil",
        "import": result,
        "sync": sync_stats,
    }), 200


@admin_bp.route('/data/export/users', methods=['GET'])
@jwt_required()
def export_users_csv():
    """Admin exports users to CSV."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    rows = User.query.order_by(User.id.asc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'id', 'name', 'email', 'account_number', 'gender', 'nik', 'address',
        'department', 'role', 'level', 'total_points', 'created_at'
    ])

    for item in rows:
        writer.writerow([
            item.id,
            item.name,
            item.email,
            item.account_number or '',
            item.gender or '',
            item.nik or '',
            item.address or '',
            item.department or '',
            item.role,
            item.level,
            int(item.total_points or 0),
            item.created_at.isoformat() if item.created_at else '',
        ])

    content = output.getvalue()
    response = current_app.response_class(content, mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=users_export.csv'
    return response


@admin_bp.route('/data/export/deposits', methods=['GET'])
@jwt_required()
def export_deposits_csv():
    """Admin exports deposits to CSV."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    rows = db.session.query(
        WasteDeposit,
        User.name.label('user_name'),
        User.account_number.label('user_account_number'),
    ).join(
        User,
        WasteDeposit.user_id == User.id,
    ).order_by(
        WasteDeposit.id.asc()
    ).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'id', 'user_id', 'user_name', 'account_number', 'weight_kg', 'waste_type',
        'status', 'points_earned', 'activity_type', 'source_waste_label',
        'source_price_per_kg', 'source_total_savings', 'created_at', 'validated_at', 'validated_by'
    ])

    for deposit, user_name, user_account_number in rows:
        writer.writerow([
            deposit.id,
            deposit.user_id,
            user_name or '',
            user_account_number or '',
            float(deposit.weight_kg or 0),
            deposit.waste_type,
            deposit.status,
            int(deposit.points_earned or 0),
            deposit.activity_type or '',
            deposit.source_waste_label or '',
            deposit.source_price_per_kg or '',
            deposit.source_total_savings or '',
            deposit.created_at.isoformat() if deposit.created_at else '',
            deposit.validated_at.isoformat() if deposit.validated_at else '',
            deposit.validated_by or '',
        ])

    content = output.getvalue()
    response = current_app.response_class(content, mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=deposits_export.csv'
    return response


@admin_bp.route('/data/reset', methods=['POST'])
@jwt_required()
def reset_member_and_deposit_data():
    """Admin resets member and transaction data while keeping admin users and master settings."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    # Delete child/member data first to satisfy FK constraints.
    deleted = {
        'user_badges': UserBadge.query.delete(synchronize_session=False),
        'user_missions': UserMission.query.delete(synchronize_session=False),
        'participation_risk': ParticipationRisk.query.delete(synchronize_session=False),
        'reward_redemptions': RewardRedemption.query.delete(synchronize_session=False),
        'waste_deposits': WasteDeposit.query.delete(synchronize_session=False),
    }

    deleted['members'] = User.query.filter(User.role == 'member').delete(synchronize_session=False)

    db.session.commit()

    return jsonify({
        "message": "Reset data anggota dan setoran berhasil",
        "deleted": {k: int(v or 0) for k, v in deleted.items()},
    }), 200


def _get_reward_upload_dir():
    upload_root = current_app.config.get('UPLOAD_FOLDER', os.path.join(os.getcwd(), 'uploads'))
    return os.path.join(upload_root, 'rewards')


def _is_allowed_image(filename, mime_type):
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in ALLOWED_IMAGE_EXTENSIONS and mime_type in ALLOWED_IMAGE_MIME_TYPES


def _remove_local_reward_image(image_url):
    if not image_url:
        return

    marker = '/api/admin/reward-images/'
    index = image_url.find(marker)
    if index == -1:
        return

    filename = image_url[index + len(marker):].strip()
    if not filename:
        return

    # Only remove flat filenames generated by this service.
    if os.path.basename(filename) != filename:
        return

    file_path = os.path.join(_get_reward_upload_dir(), filename)
    if os.path.exists(file_path):
        os.remove(file_path)


def _to_absolute_reward_image_url(image_url):
    if not image_url:
        return image_url
    if image_url.startswith('http://') or image_url.startswith('https://'):
        return image_url
    return request.host_url.rstrip('/') + image_url


def _data_dir_path():
    # api/app/routes/admin.py -> project root (TA) is parent[3]
    return os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 'data')


def _normalize_header(value: object) -> str:
    raw = str(value or '').strip()
    return ' '.join(raw.split())


def _header_map(ws) -> dict[str, int]:
    headers = [_normalize_header(c.value) for c in ws[1]]
    return {name: idx for idx, name in enumerate(headers) if name}


def _find_col(header_index: dict[str, int], aliases: list[str]) -> int | None:
    for alias in aliases:
        key = _normalize_header(alias)
        if key in header_index:
            return header_index[key]
    return None


def _detect_customer_sheet_name(wb) -> str | None:
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_index = _header_map(ws)
        account_col = _find_col(header_index, CUSTOMER_ALIASES['account_number'])
        name_col = _find_col(header_index, CUSTOMER_ALIASES['name'])
        if account_col is not None and name_col is not None:
            return sheet_name
    return None


def _clean_account_number(value: object) -> str:
    return str(value or '').strip().replace(' ', '').upper()


def _safe_text(value: object) -> str | None:
    text_value = str(value or '').strip()
    return text_value if text_value else None


def _generate_unique_import_email(account_number: str) -> str:
    base = f"{account_number.lower()}@import.dlh.local"
    if not User.query.filter_by(email=base).first():
        return base

    suffix = 1
    while True:
        candidate = f"{account_number.lower()}_{suffix}@import.dlh.local"
        if not User.query.filter_by(email=candidate).first():
            return candidate
        suffix += 1


def _dashboard_base_data(days=30):
    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    total_members = User.query.filter_by(role='member').count()
    active_members_this_month = WasteDeposit.query.filter(
        WasteDeposit.status == 'validated',
        WasteDeposit.created_at >= now.replace(day=1, hour=0, minute=0, second=0, microsecond=0),
    ).with_entities(WasteDeposit.user_id).distinct().count()

    total_deposits_today = WasteDeposit.query.filter(
        WasteDeposit.created_at >= today_start
    ).count()

    high_risk_count = ParticipationRisk.query.filter_by(risk_level='high').count()

    total_points_distributed = db.session.query(
        func.coalesce(func.sum(WasteDeposit.points_earned), 0)
    ).filter(WasteDeposit.status == 'validated').scalar()

    total_weight = db.session.query(
        func.coalesce(func.sum(WasteDeposit.weight_kg), 0)
    ).filter(WasteDeposit.status == 'validated').scalar()

    pending_deposits = WasteDeposit.query.filter_by(status='pending').count()
    pending_redemptions = RewardRedemption.query.filter_by(status='pending').count()

    days_ago = now - timedelta(days=days)
    daily_deposits = db.session.query(
        func.date(WasteDeposit.created_at).label('date'),
        func.count(WasteDeposit.id).label('count'),
        func.coalesce(func.sum(WasteDeposit.weight_kg), 0).label('weight'),
    ).filter(
        WasteDeposit.created_at >= days_ago,
    ).group_by(
        func.date(WasteDeposit.created_at)
    ).order_by(
        func.date(WasteDeposit.created_at)
    ).all()

    trend = [
        {
            'date': str(row.date),
            'deposit_count': int(row.count or 0),
            'total_weight_kg': round(float(row.weight or 0), 2),
        }
        for row in daily_deposits
    ]

    risk_distribution = db.session.query(
        ParticipationRisk.risk_level,
        func.count(ParticipationRisk.id),
    ).group_by(ParticipationRisk.risk_level).all()

    dist_map = {'low': 0, 'medium': 0, 'high': 0}
    for level, count in risk_distribution:
        if level in dist_map:
            dist_map[level] = int(count)

    recent_pending = WasteDeposit.query.filter_by(status='pending').order_by(
        WasteDeposit.created_at.desc()
    ).limit(5).all()

    kpis = {
        'total_members': int(total_members),
        'active_members_this_month': int(active_members_this_month),
        'total_deposits_today': int(total_deposits_today),
        'total_weight_kg': round(float(total_weight or 0), 2),
        'high_risk_count': int(high_risk_count),
        'total_points_distributed': int(total_points_distributed or 0),
        'pending_deposits_count': int(pending_deposits),
        'pending_redemptions_count': int(pending_redemptions),
    }

    return kpis, trend, dist_map, recent_pending


# ─── Dashboard Stats ──────────────────────────────────────────────────


@admin_bp.route('/dashboard/kpis', methods=['GET'])
@jwt_required()
def dashboard_kpis():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    kpis, _, _, _ = _dashboard_base_data()
    return jsonify({"success": True, "data": kpis}), 200


@admin_bp.route('/dashboard/trend', methods=['GET'])
@jwt_required()
def dashboard_trend():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    days = request.args.get('days', 30, type=int)
    _, trend, _, _ = _dashboard_base_data(days=days)
    return jsonify({"success": True, "data": trend}), 200


@admin_bp.route('/dashboard/risk-distribution', methods=['GET'])
@jwt_required()
def dashboard_risk_distribution():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    _, _, distribution, _ = _dashboard_base_data()
    return jsonify({"success": True, "data": distribution}), 200


@admin_bp.route('/dashboard/recent-pending', methods=['GET'])
@jwt_required()
def dashboard_recent_pending():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    _, _, _, recent_pending = _dashboard_base_data()
    return jsonify({
        "success": True,
        "data": [_serialize_deposit_with_user(item) for item in recent_pending],
    }), 200


@admin_bp.route('/deposits', methods=['GET'])
@jwt_required()
def get_admin_deposits():
    """Admin list deposits with pagination and optional filters."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '', type=str).strip()
    search = request.args.get('search', '', type=str).strip()

    query = WasteDeposit.query.join(User, WasteDeposit.user_id == User.id)

    if status:
        query = query.filter(WasteDeposit.status == status)

    if search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                User.name.ilike(like),
                User.account_number.ilike(like),
            )
        )

    pagination = query.order_by(WasteDeposit.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        "success": True,
        "data": {
            "data": [_serialize_deposit_with_user(item) for item in pagination.items],
            "total": pagination.total,
            "page": pagination.page,
            "per_page": per_page,
            "total_pages": pagination.pages,
        },
    }), 200


@admin_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def dashboard_stats():
    """Admin dashboard: KPI stats, deposit trends, risk distribution."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    # Total members
    total_members = User.query.filter_by(role='member').count()

    # Deposits today
    deposits_today = WasteDeposit.query.filter(
        WasteDeposit.created_at >= today_start
    ).count()

    # High risk count
    high_risk_count = ParticipationRisk.query.filter_by(risk_level='high').count()

    # Total points distributed (from all validated deposits)
    total_points_distributed = db.session.query(
        func.coalesce(func.sum(WasteDeposit.points_earned), 0)
    ).filter(WasteDeposit.status == 'validated').scalar()

    # Total weight collected
    total_weight = db.session.query(
        func.coalesce(func.sum(WasteDeposit.weight_kg), 0)
    ).filter(WasteDeposit.status == 'validated').scalar()

    # Pending deposits count
    pending_deposits = WasteDeposit.query.filter_by(status='pending').count()

    # Pending redemptions count
    pending_redemptions = RewardRedemption.query.filter_by(status='pending').count()

    # Deposit trend (last 30 days)
    thirty_days_ago = now - timedelta(days=30)
    daily_deposits = db.session.query(
        func.date(WasteDeposit.created_at).label('date'),
        func.count(WasteDeposit.id).label('count'),
        func.coalesce(func.sum(WasteDeposit.weight_kg), 0).label('weight'),
    ).filter(
        WasteDeposit.created_at >= thirty_days_ago,
        WasteDeposit.status == 'validated',
    ).group_by(
        func.date(WasteDeposit.created_at)
    ).order_by(
        func.date(WasteDeposit.created_at)
    ).all()

    deposit_trend = [
        {
            'date': str(row.date),
            'count': row.count,
            'weight': round(float(row.weight), 2),
        }
        for row in daily_deposits
    ]

    # Risk distribution
    risk_distribution = db.session.query(
        ParticipationRisk.risk_level,
        func.count(ParticipationRisk.id),
    ).group_by(ParticipationRisk.risk_level).all()

    risk_dist = {level: count for level, count in risk_distribution if level}

    return jsonify({
        "stats": {
            "total_members": total_members,
            "deposits_today": deposits_today,
            "high_risk_count": high_risk_count,
            "total_points_distributed": total_points_distributed,
            "total_weight_kg": round(float(total_weight), 2),
            "pending_deposits": pending_deposits,
            "pending_redemptions": pending_redemptions,
        },
        "deposit_trend": deposit_trend,
        "risk_distribution": risk_dist,
    }), 200


# ─── Member Management ────────────────────────────────────────────────

@admin_bp.route('/members', methods=['GET'])
@jwt_required()
def get_members():
    """Admin lists all members with optional search."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '').strip()
    risk_level = request.args.get('risk_level', '').strip().lower()
    sort_by = request.args.get('sort_by', 'created_at').strip().lower()

    query = User.query.filter_by(role='member')

    if search:
        query = query.filter(
            db.or_(
                User.name.ilike(f'%{search}%'),
                User.email.ilike(f'%{search}%'),
                User.account_number.ilike(f'%{search}%'),
            )
        )

    if risk_level in ('low', 'medium', 'high'):
        query = query.join(
            ParticipationRisk,
            ParticipationRisk.user_id == User.id,
        ).filter(func.lower(ParticipationRisk.risk_level) == risk_level)

    if sort_by == 'name':
        query = query.order_by(User.name.asc(), User.created_at.desc())
    elif sort_by == 'total_points':
        query = query.order_by(User.total_points.desc(), User.created_at.desc())
    else:
        query = query.order_by(User.created_at.desc())

    members = query.paginate(
        page=page, per_page=per_page, error_out=False
    )

    member_ids = [m.id for m in members.items]
    risk_map = {}
    if member_ids:
        risks = ParticipationRisk.query.filter(ParticipationRisk.user_id.in_(member_ids)).all()
        risk_map = {r.user_id: r.risk_level for r in risks}

    results = []
    for m in members.items:
        member_data = m.to_dict()
        member_data['risk_level'] = risk_map.get(m.id)
        results.append(member_data)

    return jsonify({
        "members": results,
        "total": members.total,
        "page": members.page,
        "pages": members.pages,
    }), 200


@admin_bp.route('/members', methods=['POST'])
@jwt_required()
def create_member():
    """Admin creates a new member with profile fields."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    data = request.get_json() or {}

    name = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    account_number = (data.get('account_number') or '').strip().upper() or None
    gender = (data.get('gender') or '').strip() or None
    nik = (data.get('nik') or '').strip() or None
    address = (data.get('address') or '').strip() or None
    department = (data.get('department') or '').strip() or None

    if not name or not email or not password:
        return error_response(
            "Nama, email, dan password wajib diisi",
            "validation_error",
            status=400,
            fields={
                "name": "required" if not name else None,
                "email": "required" if not email else None,
                "password": "required" if not password else None,
            },
        )

    if len(password) < 6:
        return error_response(
            "Password minimal 6 karakter",
            "validation_error",
            status=400,
            fields={"password": "min_length_6"},
        )

    if User.query.filter_by(email=email).first():
        return error_response("Email sudah digunakan", "validation_error", status=400, fields={"email": "duplicate"})

    if account_number and User.query.filter_by(account_number=account_number).first():
        return error_response("No rekening sudah digunakan", "validation_error", status=400, fields={"account_number": "duplicate"})

    if nik and User.query.filter_by(nik=nik).first():
        return error_response("NIK sudah digunakan", "validation_error", status=400, fields={"nik": "duplicate"})

    member = User(
        name=name,
        email=email,
        account_number=account_number,
        gender=gender,
        nik=nik,
        address=address,
        department=department,
        role='member',
        level='Bronze',
        total_points=0,
    )
    member.set_password(password)

    db.session.add(member)
    db.session.commit()

    return jsonify({
        "message": "Anggota berhasil ditambahkan",
        "member": member.to_dict(),
    }), 201


@admin_bp.route('/members/<int:member_id>', methods=['GET'])
@jwt_required()
def get_member_detail(member_id):
    """Admin views detailed member profile."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    member = User.query.get(member_id)
    if not member:
        return error_response("Member tidak ditemukan", "not_found", status=404)

    # Get member stats
    total_deposits = WasteDeposit.query.filter_by(user_id=member_id, status='validated').count()
    total_weight = db.session.query(
        func.coalesce(func.sum(WasteDeposit.weight_kg), 0)
    ).filter(WasteDeposit.user_id == member_id, WasteDeposit.status == 'validated').scalar()

    badges_count = UserBadge.query.filter_by(user_id=member_id).count()
    missions_completed = UserMission.query.filter_by(user_id=member_id, is_completed=True).count()

    risk = ParticipationRisk.query.filter_by(user_id=member_id).first()

    recent_deposits = WasteDeposit.query.filter_by(
        user_id=member_id
    ).order_by(WasteDeposit.created_at.desc()).limit(10).all()

    return jsonify({
        "member": member.to_dict(),
        "stats": {
            "total_deposits": total_deposits,
            "total_weight_kg": round(float(total_weight), 2),
            "badges_count": badges_count,
            "missions_completed": missions_completed,
        },
        "risk_profile": risk.to_dict() if risk else None,
        "recent_deposits": [d.to_dict() for d in recent_deposits],
    }), 200


@admin_bp.route('/members/<int:member_id>', methods=['PUT'])
@jwt_required()
def update_member_detail(member_id):
    """Admin updates basic member profile fields."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    member = User.query.get(member_id)
    if not member or member.role != 'member':
        return error_response("Member tidak ditemukan", "not_found", status=404)

    data = request.get_json() or {}

    if 'name' in data:
        member.name = (data.get('name') or '').strip()
    if 'email' in data:
        member.email = (data.get('email') or '').strip().lower()
    if 'account_number' in data:
        member.account_number = (data.get('account_number') or '').strip().upper() or None

    if 'gender' in data:
        member.gender = (data.get('gender') or '').strip() or None
    if 'nik' in data:
        member.nik = (data.get('nik') or '').strip() or None
    if 'address' in data:
        member.address = (data.get('address') or '').strip() or None
    if 'department' in data:
        member.department = (data.get('department') or '').strip() or None

    if not member.name:
        return error_response("Nama wajib diisi", "validation_error", status=400, fields={"name": "required"})
    if not member.email:
        return error_response("Email wajib diisi", "validation_error", status=400, fields={"email": "required"})

    existing_email = User.query.filter(
        User.email == member.email,
        User.id != member.id,
    ).first()
    if existing_email:
        return error_response("Email sudah digunakan", "validation_error", status=400, fields={"email": "duplicate"})

    if member.account_number:
        existing_account = User.query.filter(
            User.account_number == member.account_number,
            User.id != member.id,
        ).first()
        if existing_account:
            return error_response("No rekening sudah digunakan", "validation_error", status=400, fields={"account_number": "duplicate"})

    db.session.commit()
    return jsonify({
        "message": "Profil anggota berhasil diperbarui",
        "member": member.to_dict(),
    }), 200


@admin_bp.route('/members/<int:member_id>', methods=['DELETE'])
@jwt_required()
def delete_member(member_id):
    """Admin deletes a member and related data."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    member = User.query.get(member_id)
    if not member or member.role != 'member':
        return error_response("Member tidak ditemukan", "not_found", status=404)

    UserBadge.query.filter_by(user_id=member_id).delete(synchronize_session=False)
    UserMission.query.filter_by(user_id=member_id).delete(synchronize_session=False)
    ParticipationRisk.query.filter_by(user_id=member_id).delete(synchronize_session=False)
    RewardRedemption.query.filter_by(user_id=member_id).delete(synchronize_session=False)
    WasteDeposit.query.filter_by(user_id=member_id).delete(synchronize_session=False)

    db.session.delete(member)
    db.session.commit()

    return jsonify({"message": "Anggota berhasil dihapus"}), 200


# ─── Mission Management ───────────────────────────────────────────────

@admin_bp.route('/missions', methods=['GET'])
@jwt_required()
def get_all_missions():
    """Admin lists all missions."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    missions = Mission.query.order_by(Mission.created_at.desc()).all()
    return jsonify({"missions": [m.to_dict() for m in missions]}), 200


@admin_bp.route('/missions', methods=['POST'])
@jwt_required()
def create_mission():
    """Admin creates a new mission."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    data = request.get_json()
    if not data:
        return error_response("Request body tidak boleh kosong", "validation_error", status=400)

    title = data.get('title', '').strip()
    description = data.get('description', '').strip()
    target_type = data.get('target_type', '').strip()
    target_value = data.get('target_value')
    points_reward = data.get('points_reward')
    period = data.get('period', 'weekly').strip()

    if not title or not target_type or target_value is None or points_reward is None:
        return error_response(
            "title, target_type, target_value, dan points_reward wajib diisi",
            "validation_error",
            status=400,
            fields={
                "title": "required" if not title else None,
                "target_type": "required" if not target_type else None,
                "target_value": "required" if target_value is None else None,
                "points_reward": "required" if points_reward is None else None,
            },
        )

    if target_type not in ['deposit_count', 'weight']:
        return error_response(
            "target_type harus 'deposit_count' atau 'weight'",
            "validation_error",
            status=400,
            fields={"target_type": "invalid"},
        )

    if period not in ['daily', 'weekly']:
        return error_response("period harus 'daily' atau 'weekly'", "validation_error", status=400, fields={"period": "invalid"})

    mission = Mission(
        title=title,
        description=description,
        target_type=target_type,
        target_value=float(target_value),
        points_reward=int(points_reward),
        period=period,
        is_active=True,
    )
    db.session.add(mission)
    db.session.commit()

    return jsonify({
        "message": "Misi berhasil dibuat",
        "mission": mission.to_dict(),
    }), 201


@admin_bp.route('/missions/<int:mission_id>', methods=['PUT'])
@jwt_required()
def update_mission(mission_id):
    """Admin updates a mission."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    mission = Mission.query.get(mission_id)
    if not mission:
        return error_response("Misi tidak ditemukan", "not_found", status=404)

    data = request.get_json() or {}

    if 'title' in data:
        mission.title = data['title'].strip()
    if 'description' in data:
        mission.description = data['description'].strip()
    if 'target_type' in data:
        mission.target_type = data['target_type'].strip()
    if 'target_value' in data:
        mission.target_value = float(data['target_value'])
    if 'points_reward' in data:
        mission.points_reward = int(data['points_reward'])
    if 'period' in data:
        mission.period = data['period'].strip()
    if 'is_active' in data:
        mission.is_active = bool(data['is_active'])

    db.session.commit()

    return jsonify({
        "message": "Misi berhasil diupdate",
        "mission": mission.to_dict(),
    }), 200


# ─── Reward Management ────────────────────────────────────────────────

@admin_bp.route('/reward-images', methods=['POST'])
@jwt_required()
def upload_reward_image():
    """Admin uploads reward image with max size 5MB."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    if 'image' not in request.files:
        return error_response("File image wajib diisi", "validation_error", status=400, fields={"image": "required"})

    image = request.files['image']
    if not image or not image.filename:
        return error_response("File image tidak valid", "validation_error", status=400, fields={"image": "invalid"})

    max_size = int(current_app.config.get('MAX_REWARD_IMAGE_SIZE', 5 * 1024 * 1024))
    image.stream.seek(0, os.SEEK_END)
    image_size = image.stream.tell()
    image.stream.seek(0)
    if image_size > max_size:
        return error_response("Ukuran file melebihi batas maksimal 5MB", "payload_too_large", status=413)

    filename = secure_filename(image.filename)
    mime_type = (image.mimetype or '').lower()
    if not _is_allowed_image(filename, mime_type):
        return error_response(
            "Format file tidak didukung. Gunakan PNG, JPG, atau WEBP",
            "validation_error",
            status=400,
            fields={"image": "invalid_format"},
        )

    ext = filename.rsplit('.', 1)[1].lower()
    unique_filename = f"{uuid.uuid4().hex}.{ext}"

    upload_dir = _get_reward_upload_dir()
    os.makedirs(upload_dir, exist_ok=True)

    image.save(os.path.join(upload_dir, unique_filename))
    image_url = url_for('admin.get_reward_image', filename=unique_filename)

    return jsonify({
        "message": "Gambar reward berhasil diupload",
        "image_url": image_url,
    }), 201


@admin_bp.route('/reward-images/<path:filename>', methods=['GET'])
def get_reward_image(filename):
    """Serve uploaded reward image files."""
    return send_from_directory(_get_reward_upload_dir(), filename)

@admin_bp.route('/rewards', methods=['GET'])
@jwt_required()
def get_all_rewards():
    """Admin lists all rewards (including inactive)."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    rewards = Reward.query.order_by(Reward.created_at.desc()).all()
    payload = []
    for reward in rewards:
        item = reward.to_dict()
        item['image_url'] = _to_absolute_reward_image_url(item.get('image_url'))
        payload.append(item)

    return jsonify({"rewards": payload}), 200


@admin_bp.route('/rewards', methods=['POST'])
@jwt_required()
def create_reward():
    """Admin creates a new reward."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    data = request.get_json()
    if not data:
        return error_response("Request body tidak boleh kosong", "validation_error", status=400)

    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    points_cost = data.get('points_cost')
    stock = data.get('stock', 0)
    image_url = data.get('image_url', '').strip()

    if not name or points_cost is None:
        return error_response(
            "name dan points_cost wajib diisi",
            "validation_error",
            status=400,
            fields={
                "name": "required" if not name else None,
                "points_cost": "required" if points_cost is None else None,
            },
        )

    reward = Reward(
        name=name,
        description=description,
        points_cost=int(points_cost),
        stock=int(stock),
        image_url=image_url if image_url else None,
        is_active=True,
    )
    db.session.add(reward)
    db.session.commit()

    invalidate_cache('rewards_active')

    response_reward = reward.to_dict()
    response_reward['image_url'] = _to_absolute_reward_image_url(response_reward.get('image_url'))

    return jsonify({
        "message": "Reward berhasil dibuat",
        "reward": response_reward,
    }), 201


@admin_bp.route('/rewards/<int:reward_id>', methods=['PUT'])
@jwt_required()
def update_reward(reward_id):
    """Admin updates a reward."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    reward = Reward.query.get(reward_id)
    if not reward:
        return error_response("Reward tidak ditemukan", "not_found", status=404)

    data = request.get_json() or {}
    old_image_url = reward.image_url

    if 'name' in data:
        reward.name = data['name'].strip()
    if 'description' in data:
        reward.description = data['description'].strip()
    if 'points_cost' in data:
        reward.points_cost = int(data['points_cost'])
    if 'stock' in data:
        reward.stock = int(data['stock'])
    if 'image_url' in data:
        new_image_url = (data['image_url'] or '').strip() or None
        if old_image_url and old_image_url != new_image_url:
            _remove_local_reward_image(old_image_url)
        reward.image_url = new_image_url
    if 'is_active' in data:
        reward.is_active = bool(data['is_active'])

    db.session.commit()

    invalidate_cache('rewards_active')

    response_reward = reward.to_dict()
    response_reward['image_url'] = _to_absolute_reward_image_url(response_reward.get('image_url'))

    return jsonify({
        "message": "Reward berhasil diupdate",
        "reward": response_reward,
    }), 200


# ─── Badge Management ─────────────────────────────────────────────────

@admin_bp.route('/badges', methods=['GET'])
@jwt_required()
def get_all_badges():
    """Admin lists all badges."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    badges = Badge.query.all()
    return jsonify({"badges": [b.to_dict() for b in badges]}), 200


@admin_bp.route('/badges', methods=['POST'])
@jwt_required()
def create_badge():
    """Admin creates a new badge."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    data = request.get_json()
    if not data:
        return error_response("Request body tidak boleh kosong", "validation_error", status=400)

    name = data.get('name', '').strip()
    description = data.get('description', '').strip()
    icon_url = data.get('icon_url', '').strip()
    condition_type = data.get('condition_type', '').strip()
    condition_value = data.get('condition_value')

    if not name or not condition_type or condition_value is None:
        return error_response(
            "name, condition_type, dan condition_value wajib diisi",
            "validation_error",
            status=400,
            fields={
                "name": "required" if not name else None,
                "condition_type": "required" if not condition_type else None,
                "condition_value": "required" if condition_value is None else None,
            },
        )

    badge = Badge(
        name=name,
        description=description,
        icon_url=icon_url if icon_url else None,
        condition_type=condition_type,
        condition_value=float(condition_value),
    )
    db.session.add(badge)
    db.session.commit()

    sync_stats = sync_all_users_levels_and_badges()

    return jsonify({
        "message": "Badge berhasil dibuat",
        "badge": badge.to_dict(),
        "sync": sync_stats,
    }), 201


# ─── Reward Stock Management ──────────────────────────────────────────

@admin_bp.route('/rewards/<int:reward_id>/stock', methods=['PATCH'])
@jwt_required()
def update_reward_stock(reward_id):
    """Admin adds stock to a reward."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    reward = Reward.query.get(reward_id)
    if not reward:
        return error_response("Reward tidak ditemukan", "not_found", status=404)

    data = request.get_json() or {}
    add_stock = data.get('add_stock', 0)

    if not isinstance(add_stock, int) or add_stock <= 0:
        return error_response(
            "add_stock harus bilangan bulat positif",
            "validation_error",
            status=400,
            fields={"add_stock": "invalid"},
        )

    reward.stock += add_stock
    db.session.commit()

    invalidate_cache('rewards_active')

    return jsonify({
        "message": f"Stok berhasil ditambahkan +{add_stock}",
        "reward": reward.to_dict(),
    }), 200


# ─── Mission Soft Delete ──────────────────────────────────────────────

@admin_bp.route('/missions/<int:mission_id>', methods=['DELETE'])
@jwt_required()
def delete_mission(mission_id):
    """Admin soft-deletes a mission (set is_active=False)."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    mission = Mission.query.get(mission_id)
    if not mission:
        return error_response("Misi tidak ditemukan", "not_found", status=404)

    mission.is_active = False
    db.session.commit()

    return jsonify({
        "message": "Misi berhasil dinonaktifkan",
        "mission": mission.to_dict(),
    }), 200


# ─── Risk Trend ────────────────────────────────────────────────────────

@admin_bp.route('/risk-trend', methods=['GET'])
@jwt_required()
def risk_trend():
    """Risk distribution trend over last N months."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    months = request.args.get('months', 6, type=int)
    now = datetime.now(timezone.utc)

    results = []
    for i in range(months - 1, -1, -1):
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if i > 0:
            month_val = month_start.month - i
            year_val = month_start.year
            while month_val <= 0:
                month_val += 12
                year_val -= 1
            month_start = month_start.replace(year=year_val, month=month_val)

        # Count risk levels from participation_risk table
        risk_counts = db.session.query(
            ParticipationRisk.risk_level,
            func.count(ParticipationRisk.id),
        ).group_by(ParticipationRisk.risk_level).all()

        dist = {'low': 0, 'medium': 0, 'high': 0}
        for level, count in risk_counts:
            if level in dist:
                dist[level] = int(count)

        results.append({
            'month': month_start.strftime('%Y-%m'),
            'low': dist['low'],
            'medium': dist['medium'],
            'high': dist['high'],
        })

    return jsonify({"success": True, "data": results}), 200


# ─── Redemptions Summary ──────────────────────────────────────────────

@admin_bp.route('/redemptions/summary', methods=['GET'])
@jwt_required()
def redemptions_summary():
    """Summary of pending redemptions for admin dashboard."""
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    pending_count = RewardRedemption.query.filter_by(status='pending').count()
    total_points_held = db.session.query(
        func.coalesce(func.sum(RewardRedemption.points_spent), 0)
    ).filter(RewardRedemption.status == 'pending').scalar()

    affected_rewards = RewardRedemption.query.filter_by(
        status='pending'
    ).with_entities(RewardRedemption.reward_id).distinct().count()

    return jsonify({
        "pending_count": int(pending_count),
        "total_points_held": int(total_points_held or 0),
        "affected_rewards_count": int(affected_rewards),
    }), 200
