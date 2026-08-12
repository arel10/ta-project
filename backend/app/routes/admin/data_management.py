import csv
import io
import os
import tempfile

from flask import current_app, jsonify, request
from flask_jwt_extended import jwt_required
from openpyxl import Workbook, load_workbook

from app import db
from app.models.badge import UserBadge
from app.models.mission import UserMission
from app.models.participation_risk import ParticipationRisk
from app.models.reward import RewardRedemption
from app.models.user import User
from app.models.waste_deposit import WasteDeposit
from app.routes.admin import admin_bp
from app.routes.admin_common import (
    _clean_account_number,
    _data_dir_path,
    _detect_customer_sheet_name,
    _find_col,
    _generate_unique_import_email,
    _get_current_user,
    _header_map,
    _normalize_gender,
    _require_admin,
    _safe_text,
)
from app.services.gamification_service import sync_all_users_levels_and_badges
from app.services.simple_cache import invalidate_cache
from app.utils.api_response import error_response
from import_dlh_excel import CUSTOMER_ALIASES, import_dlh_data


@admin_bp.route('/data/import/members', methods=['POST'])
@jwt_required()
def import_members_from_file():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename:
        return error_response("File anggota wajib dipilih", "validation_error", status=400, fields={"file": "required"})

    try:
        wb = load_workbook(io.BytesIO(uploaded.read()), data_only=True)
    except Exception:
        return error_response("File anggota tidak valid. Gunakan format .xlsx", "validation_error", status=400, fields={"file": "invalid"})

    sheet_name = _detect_customer_sheet_name(wb)
    if not sheet_name:
        return error_response(
            "Sheet anggota tidak ditemukan. Pastikan ada kolom akun dan nama anggota",
            "validation_error",
            status=400,
            fields={"sheet": "missing"},
        )

    ws = wb[sheet_name]
    header_index = _header_map(ws)
    col_account = _find_col(header_index, CUSTOMER_ALIASES['account_number'])
    col_name = _find_col(header_index, CUSTOMER_ALIASES['name'])
    col_gender = _find_col(header_index, CUSTOMER_ALIASES['gender'])
    col_nik = _find_col(header_index, CUSTOMER_ALIASES['nik'])
    col_address = _find_col(header_index, CUSTOMER_ALIASES['address'])
    col_department = _find_col(header_index, CUSTOMER_ALIASES['department'])

    if col_account is None or col_name is None:
        return error_response(
            "Kolom wajib anggota tidak lengkap: No Rekening dan Nama",
            "validation_error",
            status=400,
            fields={"columns": "missing_required"},
        )

    stats = {
        'rows_seen': 0,
        'members_created': 0,
        'members_updated': 0,
        'rows_skipped': 0,
        'missing_account': 0,
        'missing_name': 0,
    }

    def _get_val(r, idx):
        if idx is None or idx < 0 or idx >= len(r):
            return None
        return r[idx]

    created_in_batch: dict[str, User] = {}

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            stats['rows_seen'] += 1

            account_number = _clean_account_number(_get_val(row, col_account))
            if not account_number:
                stats['rows_skipped'] += 1
                stats['missing_account'] += 1
                continue

            raw_name = _safe_text(_get_val(row, col_name))
            if not raw_name:
                stats['rows_skipped'] += 1
                stats['missing_name'] += 1
                continue

            name = raw_name[:64]
            gender = _normalize_gender(_get_val(row, col_gender))
            raw_nik = _safe_text(_get_val(row, col_nik))
            nik = raw_nik[:16] if raw_nik else None
            raw_address = _safe_text(_get_val(row, col_address))
            address = raw_address[:100] if raw_address else None
            raw_department = _safe_text(_get_val(row, col_department))
            department = raw_department[:64] if raw_department else None

            member = created_in_batch.get(account_number) or User.query.filter_by(account_number=account_number).first()
            if not member:
                member = User(
                    name=name,
                    email=_generate_unique_import_email(account_number),
                    account_number=account_number,
                    gender=gender,
                    nik=nik,
                    address=address,
                    department=department,
                    role='member',
                    level='Bronze',
                    total_points=0,
                )
                member.set_password('import123')
                db.session.add(member)
                created_in_batch[account_number] = member
                stats['members_created'] += 1
                continue

            updated = False
            if member.name != name:
                member.name = name
                updated = True
            if gender and member.gender != gender:
                member.gender = gender
                updated = True
            if nik and member.nik != nik:
                member.nik = nik
                updated = True
            if address and member.address != address:
                member.address = address
                updated = True
            if department and member.department != department:
                member.department = department
                updated = True

            if updated:
                stats['members_updated'] += 1

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error importing members: {e}", exc_info=True)
        return error_response(
            f"Gagal memproses import data anggota: {str(e)}",
            "import_error",
            status=400,
        )

    return jsonify({
        "message": "Import data anggota berhasil",
        "sheet": sheet_name,
        "stats": stats,
    }), 200


@admin_bp.route('/data/import/deposits', methods=['POST'])
@jwt_required()
def import_deposits_from_file():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename:
        return error_response("File setoran wajib dipilih", "validation_error", status=400, fields={"file": "required"})

    try:
        limit = int((request.form.get('limit') or '0').strip() or 0)
    except ValueError:
        return error_response("limit harus angka", "validation_error", status=400, fields={"limit": "invalid"})

    check_duplicates = str(request.form.get('check_duplicates', 'true')).lower() in ('1', 'true', 'yes', 'y')

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            uploaded.stream.seek(0)
            tmp.write(uploaded.stream.read())
            tmp_path = tmp.name

        try:
            result = import_dlh_data(
                file=tmp_path,
                customer_file=None,
                check_duplicates=check_duplicates,
                limit=max(0, limit),
                batch_size=2000,
                commit=True,
            )

            sync_stats = sync_all_users_levels_and_badges()

            return jsonify({
                "message": "Import data setoran berhasil",
                "import": result,
                "sync": sync_stats,
            }), 200
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error importing deposits: {e}", exc_info=True)
            return error_response(
                f"Gagal memproses import data setoran: {str(e)}",
                "import_error",
                status=400,
            )
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)


@admin_bp.route('/data/import', methods=['POST'])
@jwt_required()
def import_data_from_excel():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    payload = request.get_json() or {}
    check_duplicates = bool(payload.get('check_duplicates', True))
    limit = int(payload.get('limit', 0) or 0)
    batch_size = int(payload.get('batch_size', 2000) or 2000)

    data_dir = _data_dir_path()
    tx_file = os.path.join(data_dir, 'Data DLH.xlsx')
    customer_file = os.path.join(data_dir, 'data_nasabah.xlsx')

    if not os.path.exists(tx_file):
        return error_response(f"File transaksi tidak ditemukan: {tx_file}", "not_found", status=404)
    if not os.path.exists(customer_file):
        return error_response(f"File nasabah tidak ditemukan: {customer_file}", "not_found", status=404)

    result = import_dlh_data(
        file=tx_file,
        customer_file=customer_file,
        check_duplicates=check_duplicates,
        limit=limit,
        batch_size=batch_size,
        commit=True,
    )

    sync_stats = sync_all_users_levels_and_badges()

    return jsonify({
        "message": "Import data anggota dan setoran berhasil",
        "import": result,
        "sync": sync_stats,
    }), 200


@admin_bp.route('/data/export/users', methods=['GET'])
@jwt_required()
def export_users_csv():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    rows = User.query.order_by(User.id.asc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'id', 'name', 'email', 'account_number', 'gender', 'nik', 'address',
        'department', 'role', 'level', 'total_points', 'created_at'
    ])

    for item in rows:
        writer.writerow([
            item.id,
            item.name,
            item.email,
            item.account_number or '',
            item.gender or '',
            item.nik or '',
            item.address or '',
            item.department or '',
            item.role,
            item.level,
            int(item.total_points or 0),
            item.created_at.isoformat() if item.created_at else '',
        ])

    content = output.getvalue()
    response = current_app.response_class(content, mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=users_export.csv'
    return response


@admin_bp.route('/data/export/deposits', methods=['GET'])
@jwt_required()
def export_deposits_csv():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    rows = db.session.query(
        WasteDeposit,
        User.name.label('user_name'),
        User.account_number.label('user_account_number'),
    ).join(
        User,
        WasteDeposit.user_id == User.id,
    ).order_by(
        WasteDeposit.id.asc()
    ).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'id', 'user_id', 'user_name', 'account_number', 'weight_kg', 'waste_type',
        'status', 'points_earned', 'activity_type', 'source_waste_label',
        'source_price_per_kg', 'source_total_savings', 'created_at', 'validated_at', 'validated_by'
    ])

    for deposit, user_name, user_account_number in rows:
        writer.writerow([
            deposit.id,
            deposit.user_id,
            user_name or '',
            user_account_number or '',
            float(deposit.weight_kg or 0),
            deposit.waste_type,
            deposit.status,
            int(deposit.points_earned or 0),
            deposit.activity_type or '',
            deposit.source_waste_label or '',
            deposit.source_price_per_kg or '',
            deposit.source_total_savings or '',
            deposit.created_at.isoformat() if deposit.created_at else '',
            deposit.validated_at.isoformat() if deposit.validated_at else '',
            deposit.validated_by or '',
        ])

    content = output.getvalue()
    response = current_app.response_class(content, mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=deposits_export.csv'
    return response


@admin_bp.route('/data/export/deposits/xlsx', methods=['GET'])
@jwt_required()
def export_deposits_xlsx():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    rows = db.session.query(
        WasteDeposit,
        User.account_number.label('user_account_number'),
    ).join(
        User,
        WasteDeposit.user_id == User.id,
    ).order_by(
        WasteDeposit.id.asc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Setoran"

    headers = [
        'Timestamp',
        'No Rekening',
        'Jenis',
        'Berat (kg)',
        'Jenis Kegiatan',
        'Harga (Rp.) / Kg',
        'Total Tabungan (Rp.)'
    ]
    ws.append(headers)

    for deposit, user_account_number in rows:
        timestamp_str = deposit.created_at.strftime('%Y-%m-%d %H:%M:%S') if deposit.created_at else ''
        jenis_str = f"{deposit.source_waste_label or ''} ({deposit.waste_type or ''})".strip()
        if not deposit.source_waste_label and not deposit.waste_type:
            jenis_str = ''

        ws.append([
            timestamp_str,
            user_account_number or '',
            jenis_str,
            float(deposit.weight_kg or 0),
            deposit.activity_type or '',
            int(deposit.source_price_per_kg or 0) if deposit.source_price_per_kg is not None else '',
            int(deposit.source_total_savings or 0) if deposit.source_total_savings is not None else ''
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = current_app.response_class(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = 'attachment; filename=deposits_export.xlsx'
    return response


@admin_bp.route('/data/reset', methods=['POST'])
@jwt_required()
def reset_member_and_deposit_data():
    user = _get_current_user()
    admin_check = _require_admin(user)
    if admin_check:
        return admin_check

    deleted = {
        'user_badges': UserBadge.query.delete(synchronize_session=False),
        'user_missions': UserMission.query.delete(synchronize_session=False),
        'participation_risk': ParticipationRisk.query.delete(synchronize_session=False),
        'reward_redemptions': RewardRedemption.query.delete(synchronize_session=False),
        'waste_deposits': WasteDeposit.query.delete(synchronize_session=False),
    }

    deleted['members'] = User.query.filter(User.role == 'member').delete(synchronize_session=False)

    db.session.commit()

    invalidate_cache('point_settings')
    invalidate_cache('waste_point_rates_active')
    invalidate_cache('rewards_active')

    return jsonify({
        "message": "Reset data anggota dan setoran berhasil",
        "deleted": {k: int(v or 0) for k, v in deleted.items()},
    }), 200
